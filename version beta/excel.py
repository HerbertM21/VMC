import csv
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.styles import NamedStyle

# Función para agregar datos a una hoja de Excel
# Función para agregar datos a una hoja de Excel y crear un gráfico en esa hoja
def agregar_datos_a_excel(archivo_datos, archivo_existente, nombre_hoja):
    try:
        # Carga el archivo de Excel existente
        libro_excel = openpyxl.load_workbook(archivo_existente)

        # Verifica si la hoja ya existe, si no, créala
        if nombre_hoja not in libro_excel.sheetnames:
            libro_excel.create_sheet(title=nombre_hoja)

        # Selecciona la hoja por su nombre
        hoja = libro_excel[nombre_hoja]

        # Escribe los datos en la hoja de trabajo en las columnas A, B, C y D a partir de la fila especificada (fila_inicio)
        if 'A1' not in hoja:
            hoja['A1'] = 'Fecha'
        if 'B1' not in hoja:
            hoja['B1'] = 'Hora'
        if 'C1' not in hoja:
            hoja['C1'] = 'Humedad (%)'
        if 'D1' not in hoja:
            hoja['D1'] = 'Temperatura (°C)'
        if 'E1' not in hoja:
            hoja['E1'] = 'CO2 (ppm)'

        siguiente_fila = 11  # Empezar desde la fila 2 para los datos

        # Abre el archivo de datos
        with open(archivo_datos, 'r') as archivo:
            lector_csv = csv.reader(archivo)

            for fila in lector_csv:
                fecha, hora, humedad, temperatura, co2 = fila
                hoja[f'A{siguiente_fila}'] = fecha
                hoja[f'B{siguiente_fila}'] = hora
                hoja[f'C{siguiente_fila}'] = float(humedad)
                hoja[f'D{siguiente_fila}'] = float(temperatura)
                hoja[f'E{siguiente_fila}'] = float(co2)

                siguiente_fila += 1

        # Guarda el archivo de Excel con los datos agregados
        libro_excel.save(archivo_existente)

        print(f"Los datos se han agregado a '{nombre_hoja}' en {archivo_existente} correctamente.")

    except Exception as e:
        print(f"Ocurrió un error al agregar los datos a Excel: {str(e)}")

# Función para generar y mostrar el gráfico en una hoja específica
def generar_grafico(archivo_existente, nombre_hoja):
    while True:
        try:
            # Carga el archivo de Excel existente
            libro_excel = openpyxl.load_workbook(archivo_existente)
            
            # Selecciona la hoja por su nombre
            hoja = libro_excel[nombre_hoja]

            if hoja.max_row < 2:
                print("No hay suficientes datos para generar un gráfico.")
                return

            # Solicitar al usuario el rango de fechas y horas para el gráfico
            hora_inicio = input("Ingrese la hora de inicio del rango en formato HH:MM:SS: ")
            hora_fin = input("Ingrese la hora de fin del rango en formato HH:MM:SS: ")

            horas = []
            humedad = []
            temperatura = []
            co2 = []

            for fila in hoja.iter_rows(min_row=2, max_row=hoja.max_row, min_col=1, max_col=5):
                valores = [c.value for c in fila]
                if len(valores) == 5:
                    fecha, hora, h, t, c = valores
                    if hora_inicio <= hora <= hora_fin:
                        horas.append(hora)
                        humedad.append(h)
                        temperatura.append(t)
                        co2.append(c)
                else:
                    print(f"Advertencia: Se omitió una fila con datos incompletos: {valores}")



            if not horas:
                print("No hay datos disponibles para el rango de horas especificado.")
                return

            fig, ax1 = plt.subplots(figsize=(10, 6))

            ax1.set_xlabel('Hora del día')
            ax1.set_ylabel('Humedad (%)', color='tab:blue')
            ax1.plot(horas, humedad, color='tab:blue', label='Humedad (%)')
            ax1.tick_params(axis='y', labelcolor='tab:blue')

            ax2 = ax1.twinx()
            ax2.set_ylabel('Temperatura (°C)', color='tab:red')
            ax2.plot(horas, temperatura, color='tab:red', label='Temperatura (°C)')
            ax2.tick_params(axis='y', labelcolor='tab:red')

            plt.title(f'Datos en la hoja "{nombre_hoja}" para el rango de {hora_inicio} a {hora_fin}')

            fig.tight_layout()
            plt.xticks(rotation=45)
            plt.legend(loc='upper left')

            plt.show()

            generar_otro = input("¿Desea generar otro gráfico? (s/n): ")
            if generar_otro.lower() != 's':
                break

        except Exception as e:
            print(f"Ocurrió un error al generar el gráfico: {str(e)}")